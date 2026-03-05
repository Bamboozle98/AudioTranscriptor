import csv
from datetime import datetime
from typing import Any, Dict, List, Optional

try:
    from openpyxl import Workbook, load_workbook
except ImportError:
    Workbook = None
    load_workbook = None


FIELDNAMES = ["session_id", "timestamp", "item", "amount", "unit", "amount_oz"]


SYSTEM_RECORD_PARSER = """
You are extracting restaurant inventory measurements from speech-to-text.

Return ONLY valid JSON (no markdown, no commentary).
Extract ALL records mentioned.

Output a JSON array like:
[
  {"item": "string", "amount": number, "unit": "oz|lb|count"}
]

Rules:
- Extract every occurrence of: <item phrase> + <amount> + <unit>.
- Units:
  - ounce/ounces/oz -> "oz"
  - pound/pounds/lb/lbs -> "lb"
- If the user gives a number with NO unit and it is clearly a quantity count (e.g., "pepsi 3", "2oz cups 36"):
  use unit="count".
- Treat comma-split phrases as a single item phrase (e.g., "chipotle, aioli, 20 ounces" -> item="chipotle aioli").
- Convert number words to numerals when needed ("one" -> 1).
- Do not invent items or amounts.
- If an item has no clear number, omit it.
"""


_UNIT_ALIASES = {
    "oz": "oz", "ounce": "oz", "ounces": "oz",
    "lb": "lb", "lbs": "lb", "pound": "lb", "pounds": "lb",
    "count": "count",
}

def _to_float(x: Any) -> float:
    if isinstance(x, (int, float)):
        return float(x)
    if isinstance(x, str):
        return float(x.strip())
    raise ValueError("amount must be a number")

def _norm_unit(u: Any) -> str:
    u = "" if u is None else str(u).strip().lower()
    if u in _UNIT_ALIASES:
        return _UNIT_ALIASES[u]
    raise ValueError(f"unit must be oz, lb, or count (got: {u})")

def _to_ounces(amount: float, unit: str) -> Optional[float]:
    if unit == "oz":
        return amount
    if unit == "lb":
        return amount * 16.0
    return None  # count has no ounces


def validate_and_normalize_record(rec: Dict[str, Any], session_id: str) -> Dict[str, Any]:
    for k in ["item", "amount", "unit"]:
        if k not in rec:
            raise ValueError(f"Missing key: {k}")

    item = (rec["item"] or "").strip()
    if not item:
        raise ValueError("item cannot be empty")

    amount = _to_float(rec["amount"])
    if amount <= 0:
        raise ValueError("amount must be > 0")

    unit = _norm_unit(rec["unit"])

    # If count, force integer
    if unit == "count":
        amount = int(round(amount))

    amount_oz = _to_ounces(amount, unit)
    amount_oz_out = "" if amount_oz is None else round(amount_oz, 3)

    return {
        "session_id": session_id,
        "timestamp": datetime.now().isoformat(timespec="seconds"),
        "item": item,
        "amount": amount,
        "unit": unit,
        "amount_oz": amount_oz_out,
    }


def append_records_to_csv(csv_path: str, records: List[Dict[str, Any]]) -> None:
    file_exists = False
    try:
        with open(csv_path, "r", newline="", encoding="utf-8") as _:
            file_exists = True
    except FileNotFoundError:
        file_exists = False

    with open(csv_path, "a", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=FIELDNAMES)
        if not file_exists:
            w.writeheader()
        for rec in records:
            w.writerow({k: rec.get(k, "") for k in FIELDNAMES})


def append_records_to_xlsx(xlsx_path: str, records: List[Dict[str, Any]], sheet_name: str = "weigh_ins") -> None:
    if Workbook is None or load_workbook is None:
        raise ImportError("openpyxl is required for XLSX support. Install with: pip install openpyxl")

    try:
        wb = load_workbook(xlsx_path)
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.create_sheet(sheet_name)
    except FileNotFoundError:
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        ws.append(FIELDNAMES)

    # Ensure header exists
    if ws.max_row < 1 or ws.cell(row=1, column=1).value != FIELDNAMES[0]:
        ws.delete_rows(1, ws.max_row)
        ws.append(FIELDNAMES)

    for rec in records:
        ws.append([rec.get(k, "") for k in FIELDNAMES])

    wb.save(xlsx_path)